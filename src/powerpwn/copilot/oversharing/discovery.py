import asyncio
import os
import re

import openpyxl

from powerpwn.copilot.copilot_connector.copilot_connector import CopilotConnector
from powerpwn.copilot.models.chat_argument import ChatArguments


class Discovery:
    def __init__(self, parsed_args: ChatArguments, prompts_file="pii.txt", output_file="oversharedfiles_report1.xlsx"):
        self.prompts_file = prompts_file
        self.output_file = output_file
        self.copilot_connector = CopilotConnector(parsed_args)
        self.prompts = self.read_prompts_from_file(self.prompts_file)

    def read_prompts_from_file(self, file_path):
        """
        Reads prompts from the file. Expects sections separated by two newlines.
        """
        prompts = {}
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read().strip()

            # Split on double newlines
            sections = content.split("\n\n")
            for section in sections:
                lines = section.split("\n", 1)
                if len(lines) == 2:
                    key = lines[0].strip(":").strip()
                    value = lines[1].strip()
                    prompts[key] = value
                else:
                    print(f"Warning: Skipping invalid section: {section}")
        return prompts

    def enhanced_parser(self, raw_message):
        """
        Parses the LLM response in a chunk-by-chunk manner to extract file info.
        """
        # Split by numbered items or bullet points
        chunks = re.split(r"\n\s*(?=\d+\.\s)", raw_message)

        # A small set of known file extensions we might expect.
        # You can expand as needed.
        FILE_EXTENSIONS = [".docx", ".xlsx", ".csv", ".pdf", ".pptx", ".aspx"]

        files = []

        for chunk in chunks:
            file_info = {}

            # Option 1: If there's a "File Name:" or "File:" pattern, handle as before
            match_file = re.search(r"(?:File Name\s*:|File\s*:|File:\s*)(.*)", chunk, re.IGNORECASE)

            # We will store a "candidate filename" in file_info["file_name"]
            # and "candidate link" in file_info["file_link"]
            if match_file:
                file_name_line = match_file.group(1).strip()
                # Remove trailing bracket references like [1], [2], etc., if any
                file_name_line = re.sub(r"\[\d+\]", "", file_name_line).strip()

                # Check for Markdown link pattern: [someFile.ext](URL)
                bracket_match = re.search(r"\[([^\]]+)\]\(([^)]+)\)", file_name_line)
                if bracket_match:
                    file_info["file_name"] = bracket_match.group(1).strip()
                    file_info["file_link"] = bracket_match.group(2).strip()
                else:
                    # Fallback if there's no bracket/link pattern
                    file_info["file_name"] = file_name_line
            else:
                # No explicit "File:" line found => fallback to extension-based detection
                # We look for something that ends with .docx or .csv, etc.
                # Then, in the same chunk, see if there's a URL (http/https).
                # This might look for lines like:
                # "employee_records.csv[1](https://...)"
                # or "SensitiveInformation.docx (http://...)"

                # a) Look for any potential link
                url_match = re.search(r"(https?://[^\s]+)", chunk)
                if url_match:
                    candidate_url = url_match.group(1).strip()
                    file_info["file_link"] = candidate_url
                else:
                    candidate_url = None  # no link found

                # b) Try to find a filename with known extension before or around that link
                #    We can attempt a regex that captures something like "SomeFile.pdf[1]" or "SomeFile.pdf"
                #    prior to the link.
                if candidate_url:
                    # If there's a bracket pattern, e.g. "employee_records.csv[1](...)"
                    # we can capture everything up to "[1]("
                    # Example: "employee_records.csv[1](https://...)"
                    # We'll match group(1) = "employee_records.csv"
                    pattern_before_url = rf"([^\s]+(?:{'|'.join(FILE_EXTENSIONS)})(?:\[\d+\])?)\("
                    fallback_match = re.search(pattern_before_url, chunk, re.IGNORECASE)
                    if fallback_match:
                        # strip any trailing [1], [2], etc.
                        raw_name = re.sub(r"\[\d+\]$", "", fallback_match.group(1).strip())
                        file_info["file_name"] = raw_name
                    else:
                        # If we can't find that pattern, look for a simpler approach
                        # e.g. "employee_records.csv https://..."
                        # We'll assume there's a space or punctuation before the link
                        pattern_before_url = rf"([^\s]+(?:{'|'.join(FILE_EXTENSIONS)}))(?=\s*\(?https?)"
                        fallback_match2 = re.search(pattern_before_url, chunk, re.IGNORECASE)
                        if fallback_match2:
                            file_info["file_name"] = fallback_match2.group(1).strip()

            # 2) Extract the author line or “Author:”
            #    e.g., "- **Author:** Kris Smith[1](...)"
            #    or "The author is indicated as Kris Smith"
            match_author = re.search(r"(?:Author[s]*:\s*|The\s+author[s]?\s+(?:is|are)\s+(?:indicated\s+as\s+)?)([^\n]+)", chunk, re.IGNORECASE)
            if match_author:
                authors = match_author.group(1).strip()
                # Remove markdown asterisks
                authors = authors.replace("**", "")
                # Remove bracket references like [1](http...)
                authors = re.sub(r"\[\d+\]\([^)]*\)", "", authors).strip()
                file_info["author"] = authors

            # If we also want "Contains:" or "File Type:" or "Last Modified:"
            match_file_type = re.search(r"(?:File Type:\s*)([^\n]+)", chunk, re.IGNORECASE)
            if match_file_type:
                file_info["file_type"] = match_file_type.group(1).strip()

            match_contains = re.search(r"(?:Contains:\s*)([^\n]+)", chunk, re.IGNORECASE)
            if match_contains:
                file_info["contains"] = [c.strip() for c in match_contains.group(1).split(",")]

            # 4) Validate file_info
            if "file_name" in file_info or "file_link" in file_info:
                file_info.setdefault("file_name", "N/A")
                file_info.setdefault("file_link", "N/A")
                file_info.setdefault("author", "N/A")

                # Optionally filter out if it doesn't look legit
                # (for example, if file_name is still "1" or "N/A")
                valid_ext = any(file_info["file_name"].lower().endswith(ext) for ext in FILE_EXTENSIONS)
                # If you want to ensure there's at least a link or something:
                if valid_ext or file_info["file_link"].startswith("http"):
                    files.append(file_info)

        return files

    def save_to_excel(self, categorized_files):
        """
        Append new file entries to 'pii_sensitive_files_report.xlsx' if not duplicates.
        Each file will have its own separate columns for File Name, File Link, and Author.
        """
        if os.path.exists(self.output_file):
            wb = openpyxl.load_workbook(self.output_file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sensitive PII Files"
            headers = ["File Name", "File Link", "Author", "Contains"]
            ws.append(headers)

        existing_files = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_files.add((row[0], row[1]))  # Check for duplicates based on file name and link

        for file in categorized_files:
            file_name = file["file_name"]
            file_link = file["file_link"]
            author = file["author"]
            if (file_name, file_link) not in existing_files:
                # Append separate file details to the Excel sheet
                ws.append([file_name, file_link, author, ", ".join(file.get("contains", []))])
                existing_files.add((file_name, file_link))  # Add to existing files set to avoid duplicates

        wb.save(self.output_file)
        print(f"Excel file updated: {self.output_file}")

    def post_process_files(self, files_list):
        """
        Given a list of dictionaries (each with file_name, file_link, author, etc.),
        perform some clean-up:
        1) Remove duplicates by file link.
        2) Skip files whose file_name is numeric or N/A.
        3) Clean author field by removing bracket references and parentheses text.
        """

        unique_links = set()
        cleaned_files = []

        for f in files_list:
            link = f.get("file_link", "")

            # 1) Remove duplicates based on link
            #    If the link is already encountered, skip
            if link in unique_links:
                continue
            unique_links.add(link)

            # 2) Skip if file_name is numeric (e.g. '1') or 'N/A'
            file_name = f.get("file_name", "").strip()
            if file_name.isdigit() or file_name == "N/A":
                continue

            # 3) Clean up author field
            #    Remove everything in parentheses ( ... ) and also [ ... ](...) references
            #    e.g. "[Kris Smith](https://...)" or "Kris Smith (Dept Head)"
            author = f.get("author", "")

            # Remove "[anything](anything)" references first
            author = re.sub(r"\[.*?\]\(.*?\)", "", author)

            # Remove parentheses and their contents, e.g. "(Dept Head)"
            author = re.sub(r"\(.*?\)", "", author)

            # Trim whitespace, semicolons, extra commas
            author = author.strip().strip(";").strip(",")

            # If you'd like to replace semicolons with commas (in case of multiple authors)
            # author = author.replace(";", ",").strip(",")

            # Put cleaned author back into the dictionary
            f["author"] = author if author else "N/A"

            # Add the cleaned file entry to the final list
            cleaned_files.append(f)

        return cleaned_files

    def handle_response(self, raw_message):
        """
        Processes the raw response using the enhanced parser.
        """
        # Check if raw_message is a WebsocketParsedMessage object
        if hasattr(raw_message, "copilot_message"):
            raw_message = raw_message.copilot_message  # Extract the actual message

        if not isinstance(raw_message, str):
            raise TypeError("raw_message must be a string or an object containing a message string.")

        extracted_files = self.enhanced_parser(raw_message)
        print("DEBUG: Extracted files:", extracted_files)
        return extracted_files

    async def process_prompt(self, prompt):
        """
        Sends a prompt to Copilot and handles the response.
        """
        print(f"Sending prompt: {prompt}")
        result = await asyncio.gather(self.copilot_connector.connect(prompt))

        if result and result[0]:
            raw_message = result[0].parsed_message
            print(f"Raw Response: {raw_message}")
            files_list = self.handle_response(raw_message)
            if files_list:
                self.save_to_excel(files_list)

    def run(self):
        """
        The main method to run the entire discovery process.
        """
        # Initialize connection to Copilot
        self.copilot_connector.init_connection()

        # Process initial and second prompts
        asyncio.run(self.process_prompt(self.prompts["init_prompt"]))
        asyncio.run(self.process_prompt(self.prompts["second_prompt"]))

        # Dynamic third prompts
        pii_types = [
            "*Social Security Numbers*",
            "*Employee Records*",
            "*Employee Records* such as start dates, end dates, and date of births",
            "*Employee Records* such as salary and job position",
            "*Employee Records* such as addresses",
            "*Passport Numbers*",
            "*Driver's License Numbers*",
            "*Employee Salaries*",
            "*IP addresses*",
            "*Contact Information such as phone numbers*",
            "*Contact Information such as emails*",
        ]

        for pii_type in pii_types:
            third_prompt_dynamic = f"great job! but are there any other sensitive files that contains {pii_type}? i believe you are missing quite a lot. For each file be sure to include the file name, file link and author. be sure to also search for different formats such as .csv, .xlsx, and .pdf!"
            print(f"Sending third prompt for PII type: {pii_type}")
            asyncio.run(self.process_prompt(third_prompt_dynamic))

        print("Finished processing.")


# Example usage:
if __name__ == "__main__":
    discovery = Discovery()
    discovery.run()